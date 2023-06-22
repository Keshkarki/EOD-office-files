
""" Making py file Notes:
1. Experimental stage
2. have structure excel already just need to fil data daily
3. if file not present for today will print file not present for today's date
    """


# %%
#--> importing libraries
import numpy as np
import pandas as pd
import datetime as dt
import openpyxl
import os

pd.set_option('display.max_columns', None)


# %% geting report file accessing it first taking same day file only

filesList = os.listdir("C:\\keshav\\dailyReport\\dailyFiles2")
filesList

# today's date
today = dt.date.today()

# will open today's file only

todayFileFound = False

for file in filesList:
    datestr = file.split('.')[0]
    datedt = pd.to_datetime(datestr).date()
    if datedt == today:
        df = pd.read_csv(f"C:\\keshav\\dailyReport\\dailyFiles2\\{file}")

        todayFileFound = True

        break

# Check if today's file is found or not
if todayFileFound:
    pass

else:
    print("Today's file is not present.")












# %% code fro extracting all vaues from dataframe

df['date'] = today
df.set_index('date',inplace=True)

#adding symbol column with help of exchange symbol
df.insert(5,'Symbol',df['Exchange Symbol'].str.split().str.get(0))

# filtering only for NIFTY,BANKNIFTY,FINNIFTY related to underlaying our dictionary
df = df[df['Symbol'].isin(['NIFTY','BANKNIFTY','FINNIFTY'])]


#calculating lot size column with help of classes b

from packages.underlaying import lot_size
df.insert(11,'lot_size',df['Symbol'].apply(lot_size))
df.insert(13,'no_of_lots',(df['Qty']/df['lot_size']).astype(int))


# calculating brokerage with help of class
from packages.brokerage import brokerage
df.insert(14,'brokerage_per_lot',df['Symbol'].apply(brokerage))
df.insert(15,'brokerage_cost',df['no_of_lots']*df['brokerage_per_lot'])

## importing additional cost
from packages.additionalCost import additional_cost
df.insert(16,'additional_cost',(df['Symbol'].apply(additional_cost)*df['brokerage_cost']).astype(int))
df.insert(17,'Total_cost',(df['brokerage_cost']+df['additional_cost']))


#calculating Txn
def new_price(row):
    if row['Txn'] =='SELL':
        return (row['Avg Price']*row['Qty']-row['Total_cost'])/row['Qty']
    
    else:
        return (row['Avg Price']*row['Qty']+row['Total_cost'])/row['Qty']
df['New Price']= df.apply(new_price,axis=1)


def new_amount(row):
    if row['Txn'] == 'SELL':
        return row['New Price']*row['Qty']
    
    else:
        return -(row['New Price']*row['Qty'])

df['New Amount']=df.apply(new_amount,axis=1)


df.insert(2,'P_names',df['Portfolio Name'].str.extract('(CSL_STRANGS|70PCT|CP_2RE|25PCT_2RE)'))
grp = df.groupby(['User Alias','P_names'])[['New Price','New Amount']].sum().reset_index()

grp['date'] =today
grp.set_index('date',inplace=True)

















# %% MAKING PIVOT

#clients Total Amount as per date  ****************** 
clientAmountsPivot = grp.pivot_table(index='date', columns='User Alias', values='New Amount', aggfunc='sum')

# Strategies as per date
strategies = grp.pivot_table(index='P_names', columns='User Alias', values='New Amount', aggfunc='sum').reset_index()
strategies['date'] =today
strategies.set_index('date',inplace=True)










# %%  # --EXCEL READ WRITE
# loading worksheet and checkng value accessible or not
wb = openpyxl.load_workbook('C:\\keshav\\dailyReport\\StoxxoDailyReport2.xlsx')


#   loading all sheets  
ws1 = wb['clientReport']
ws2 = wb['25PCT_2RE']
ws3 = wb['70PCT']
ws4 = wb['CP_2RE']
ws5 = wb['CSL_STRANGS']

# ws5['C3'].value
ws1.cell(row=3,column=2).value




i = len(filesList)  #row accessor for new file if 


#Filling values in ws1 clientReport sheet 
date = clientAmountsPivot.index[0].strftime('%d-%m-%Y')

try:
    ws1.cell(row=i+1,column=1).value = date

except:
    pass


try:
    ws1.cell(row=i+1,column=2).value = clientAmountsPivot['AD1_IIFL'].values[0]

except:
    pass


try:
    ws1.cell(row=i+1,column=3).value = clientAmountsPivot['NMEDEL'].values[0]

except:
    pass


try:
    ws1.cell(row=i+1,column=4).value = clientAmountsPivot['RKEDEL'].values[0]

except:
    pass


try:
    ws1.cell(row=i+1,column=5).value = clientAmountsPivot['SS3POSIIFL'].values[0]

except:
    pass


try:
    ws1.cell(row=i+1,column=6).value = clientAmountsPivot['SIMULATED1'].values[0]

except:
    pass


# %% Making sepearte rows for each strategy
stg1 = strategies[strategies['P_names'] == '25PCT_2RE']
stg2 = strategies[strategies['P_names'] == '70PCT']
stg3 = strategies[strategies['P_names'] == 'CP_2RE']
stg4 = strategies[strategies['P_names'] == 'CSL_STRANGS']



# %% for sheet 2 25PCT_2RE

#Filling values in ws2 clientReport sheet 
date = clientAmountsPivot.index[0].strftime('%d-%m-%Y')

try:
    ws2.cell(row=i+1,column=1).value = date

except:
    pass


try:
    ws2.cell(row=i+1,column=2).value = stg1['P_names'][0]

except:
    pass


try:
    ws2.cell(row=i+1,column=3).value = stg1['AD1_IIFL'][0]

except:
    pass


try:
    ws2.cell(row=i+1,column=4).value = stg1['NMEDEL'][0]

except:
    pass


try:
    ws2.cell(row=i+1,column=5).value = stg1['RKEDEL'][0]

except:
    pass


try:
    ws2.cell(row=i+1,column=6).value = stg1['SS3POSIIFL'][0]

except:
    pass




# %% for sheet 3 25PCT_2RE

#Filling values in ws3 clientReport sheet 
date = clientAmountsPivot.index[0].strftime('%d-%m-%Y')

try:
    ws3.cell(row=i+1,column=1).value = date

except:
    pass


try:
    ws3.cell(row=i+1,column=2).value = stg2['P_names'][0]

except:
    pass


try:
    ws3.cell(row=i+1,column=3).value = stg2['AD1_IIFL'][0]

except:
    pass


try:
    ws3.cell(row=i+1,column=4).value = stg2['NMEDEL'][0]

except:
    pass


try:
    ws3.cell(row=i+1,column=5).value = stg2['RKEDEL'][0]

except:
    pass


try:
    ws3.cell(row=i+1,column=6).value = stg2['SS3POSIIFL'][0]

except:
    pass




# %% for sheet 4 cp_2RE

#Filling values in ws4 clientReport sheet 
date = clientAmountsPivot.index[0].strftime('%d-%m-%Y')

try:
    ws4.cell(row=i+1,column=1).value = date

except:
    pass


try:
    ws4.cell(row=i+1,column=2).value = stg3['P_names'][0]

except:
    pass


try:
    ws4.cell(row=i+1,column=3).value = stg3['AD1_IIFL'][0]

except:
    pass


try:
    ws4.cell(row=i+1,column=4).value = stg3['NMEDEL'][0]

except:
    pass


try:
    ws4.cell(row=i+1,column=5).value = stg3['RKEDEL'][0]
except:
    pass


try:
    ws4.cell(row=i+1,column=6).value = stg3['SS3POSIIFL'][0]

except:
    pass



# %% for sheet 5 CSL_strangs

#Filling values in ws5 clientReport sheet 
date = clientAmountsPivot.index[0].strftime('%d-%m-%Y')

try:
    ws5.cell(row=i+1,column=1).value = date

except:
    pass


try:
    ws5.cell(row=i+1,column=2).value = stg4['P_names'][0]

except:
    pass


try:
    ws5.cell(row=i+1,column=3).value = stg4['AD1_IIFL'][0]

except:
    pass


try:
    ws5.cell(row=i+1,column=4).value = stg4['NMEDEL'][0]

except:
    pass


try:
    ws5.cell(row=i+1,column=5).value = stg4['RKEDEL'][0]

except:
    pass


try:
    ws5.cell(row=i+1,column=6).value = stg4['SS3POSIIFL'][0]

except:
    pass

# %% saving to excel
wb.save('C:\\keshav\\dailyReport\\StoxxoDailyReport2.xlsx')



