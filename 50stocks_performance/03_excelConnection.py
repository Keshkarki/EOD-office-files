""" PART 3 EXCEL CONNECTION"""
#Note --> exe file not working with master df reading hence making 04 for this
# %%  (1) # importing libraries

import numpy as np
import pandas as pd
import yfinance as yf
from datetime import timedelta
pd.set_option('display.max_columns', None)
pd.set_option('display.width', 300)

import warnings
# Suppress the DeprecationWarning
warnings.filterwarnings("ignore", category=DeprecationWarning)



# %%  (2) #reading file
df = pd.read_csv('C:\\keshav\\50stocks_performance\\masterFile.csv')
# print(df)

df['Date'] = pd.to_datetime(df['Date'], format='%d-%m-%Y')
df.set_index('Date',inplace=True)
df


# %%  (3)  #making empty dataframe
companyLiist = df.columns.to_list()

index_columns = companyLiist
empty_columns = ['daily', 'weekly', 'monthly', 'quarterly', 'yearly', 'ytd', 'custom']
main_df = pd.DataFrame(index=index_columns, columns=empty_columns)
main_df


# %%  (4) ---->INPUT DATE

import openpyxl
wb = openpyxl.load_workbook('C:\\keshav\\50stocks_performance\\mainFile.xlsx')
ws =  wb['Sheet2']   #note #imp #change sheet




# inputDate = "2021-08-12"
# inputDateDt = pd.to_datetime(inputDate)
# print(inputDateDt)
# print(type(inputDateDt))

# fromDate = '2021-08-11'
# fromDateDt = pd.to_datetime(fromDate)
# print(fromDateDt)
# print(type(fromDateDt))




inputDateDt = ws.cell(row=2,column=2).value
inputDateDt = pd.Timestamp(inputDateDt)
print(inputDateDt)
print(type(inputDateDt))


fromDateDt = ws.cell(row=2,column=5).value
fromDateDt = pd.Timestamp(fromDateDt)
print(fromDateDt)
print(type(fromDateDt))



# dates setup for ytd
year = inputDateDt.year
prevYear = year-1

dateString = f"31-march-{year}"
dateString2 = f"31-march-{prevYear}"


x = pd.to_datetime(dateString)
y = pd.to_datetime(dateString2)



# (5) --> MAIN CODE
if inputDateDt in df.index:
    currentPriceList = df.loc[inputDateDt]
else:
    print(f'{inputDateDt} not present in data take another date as current date')


# 1 day before     (daily)
dailyBaseDate =  inputDateDt - timedelta(days=1)

#   7 days before v (weekly)
weeklyBaseDate = inputDateDt - timedelta(days=7)

# 30 days before  (monthly)
monthlyBaseDate = inputDateDt - timedelta(days=30)


#90 days before (quarterly)
quarterlyBaseDate = inputDateDt - timedelta(days=90)

# 365 days before (yearly)
yearlyBaseDate = inputDateDt - timedelta(days=365)

#ytd 
if inputDateDt>x:
    ytdBaseDate = x

else:
    ytdBaseDate = y


#FROM PARTICULAR BASE DATE
customBaseDate = fromDateDt


## daily
if dailyBaseDate >= df.index[0] and dailyBaseDate <= df.index[-1]:  #if base date in range
    if dailyBaseDate in df.index:                                             #if base date present directly
        basePriceList  = df.loc[dailyBaseDate]
        dailyReturnPer =  round(((currentPriceList/basePriceList -1)*100),1).values
        main_df.loc[:,'daily'] = dailyReturnPer


    else:
        temp_df = df[df.index < dailyBaseDate]  
        nearest_index = temp_df.index.searchsorted(inputDateDt)-1      #will get index
        date = temp_df.index[nearest_index]                                         #will get date 
        basePriceList = temp_df.loc[date]                                           #will get row or values
        dailyReturnPer =  round(((currentPriceList/basePriceList -1)*100),1).values
        main_df.loc[:,'daily'] = dailyReturnPer
else:
    print('Date is out of range of dates')


## weekly
if weeklyBaseDate >= df.index[0] and weeklyBaseDate <= df.index[-1]:  #if base date in range
    if weeklyBaseDate in df.index:                                             #if base date present directly
        basePriceList  = df.loc[weeklyBaseDate]
        weeklyReturnPer =  round(((currentPriceList/basePriceList -1)*100),1).values
        main_df.loc[:,'weekly'] = weeklyReturnPer


    else:
        temp_df = df[df.index < weeklyBaseDate]  
        nearest_index = temp_df.index.searchsorted(inputDateDt)-1      #will get index
        date = temp_df.index[nearest_index]                                         #will get date 
        basePriceList = temp_df.loc[date]                                           #will get row or values
        weeklyReturnPer =  round(((currentPriceList/basePriceList -1)*100),1).values
        main_df.loc[:,'weekly'] = weeklyReturnPer
else:
    print('Date is out of range of dates')



## monthly
if monthlyBaseDate >= df.index[0] and monthlyBaseDate <= df.index[-1]:  #if base date in range
    if monthlyBaseDate in df.index:                                             #if base date present directly
        basePriceList  = df.loc[monthlyBaseDate]
        monthlyReturnPer =  round(((currentPriceList/basePriceList -1)*100),1).values
        main_df.loc[:,'monthly'] = monthlyReturnPer


    else:
        temp_df = df[df.index < monthlyBaseDate]  
        nearest_index = temp_df.index.searchsorted(inputDateDt)-1      #will get index
        date = temp_df.index[nearest_index]                                         #will get date 
        basePriceList = temp_df.loc[date]                                           #will get row or values
        monthlyReturnPer =  round(((currentPriceList/basePriceList -1)*100),1).values
        main_df.loc[:,'monthly'] = monthlyReturnPer
else:
    print('Date is out of range of dates')



## quarterly
if quarterlyBaseDate >= df.index[0] and quarterlyBaseDate <= df.index[-1]:  #if base date in range
    if quarterlyBaseDate in df.index:                                             #if base date present directly
        basePriceList  = df.loc[quarterlyBaseDate]
        quarterlyReturnPer =  round(((currentPriceList/basePriceList -1)*100),1).values
        main_df.loc[:,'quarterly'] = quarterlyReturnPer


    else:
        temp_df = df[df.index < quarterlyBaseDate]  
        nearest_index = temp_df.index.searchsorted(inputDateDt)-1      #will get index
        date = temp_df.index[nearest_index]                                         #will get date 
        basePriceList = temp_df.loc[date]                                           #will get row or values
        quarterlyReturnPer =  round(((currentPriceList/basePriceList -1)*100),1).values
        main_df.loc[:,'quarterly'] = quarterlyReturnPer
else:
    print('Date is out of range of dates')



## yearly
if yearlyBaseDate >= df.index[0] and yearlyBaseDate <= df.index[-1]:  #if base date in range
    if yearlyBaseDate in df.index:                                             #if base date present directly
        basePriceList  = df.loc[yearlyBaseDate]
        yearlyReturnPer =  round(((currentPriceList/basePriceList -1)*100),1).values
        main_df.loc[:,'yearly'] = yearlyReturnPer


    else:
        temp_df = df[df.index < yearlyBaseDate]  
        nearest_index = temp_df.index.searchsorted(inputDateDt)-1      #will get index
        date = temp_df.index[nearest_index]                                         #will get date 
        basePriceList = temp_df.loc[date]                                           #will get row or values
        yearlyReturnPer =  round(((currentPriceList/basePriceList -1),1)*100).values
        main_df.loc[:,'yearly'] = yearlyReturnPer
else:
    print('Date is out of range of dates')



## ytd
if ytdBaseDate >= df.index[0] and ytdBaseDate <= df.index[-1]:  #if base date in range
    if ytdBaseDate in df.index:                                             #if base date present directly
        basePriceList  = df.loc[ytdBaseDate]
        ytdReturnPer =  round(((currentPriceList/basePriceList -1)*100),1).values
        main_df.loc[:,'ytd'] = ytdReturnPer


    else:
        temp_df = df[df.index < ytdBaseDate]  
        nearest_index = temp_df.index.searchsorted(inputDateDt)-1      #will get index
        date = temp_df.index[nearest_index]                                         #will get date 
        basePriceList = temp_df.loc[date]                                           #will get row or values
        ytdReturnPer =  round(((currentPriceList/basePriceList -1)*100),1).values
        main_df.loc[:,'ytd'] = yearlyReturnPer
else:
    print('Date is out of range of dates')


#custom from date
if customBaseDate >= df.index[0] and customBaseDate <= df.index[-1]:  #if base date in range
    if customBaseDate in df.index:                                             #if base date present directly
        basePriceList  = df.loc[customBaseDate]
        customReturnPer =  round(((currentPriceList/basePriceList -1)*100),1).values
        main_df.loc[:,'custom'] = customReturnPer


    else:
        temp_df = df[df.index < customBaseDate]  
        nearest_index = temp_df.index.searchsorted(inputDateDt)-1      #will get index
        date = temp_df.index[nearest_index]                                         #will get date 
        basePriceList = temp_df.loc[date]                                           #will get row or values
        customReturnPer =  round(((currentPriceList/basePriceList -1)*100),1).values
        main_df.loc[:,'custom'] = yearlyReturnPer
else:
    print('Date is out of range of dates')


# print(main_df)




# ( 6 ) --> SPLIT NIFTY AND CONNECTING WITH EXCEL

other_df = main_df[main_df.index != 'Nifty']
nifty_df = main_df[main_df.index == 'Nifty']


# 




#getting nifty values as a list and updating nifty row
nifty_list = nifty_df.values[0]
for i in range(2,9):
    ws.cell(row = 1,column=i).value = nifty_list[i-2]



#first checking if companies is in ascending order or not
if ws.cell(row=4, column=1).value == 'ADANIENT':
    if ws.cell(row=53, column=1).value == 'WIPRO':

        for i in range(4,54):
            for j in range(2,9):
                ws.cell(row = i,column = j).value = other_df.iloc[i-4][j-2]

       
else:
    print('!!!!!!!!!!!FIRST SHORT A-Z Adani should be in top else result will be False')


wb.save('C:\\keshav\\50stocks_performance\\mainFile.xlsx')

print('Done')


# %%
